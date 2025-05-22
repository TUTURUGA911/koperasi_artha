from flask import Flask, request, jsonify, render_template, redirect, url_for, send_file
from pymongo import MongoClient, ASCENDING
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from bson import ObjectId
from functools import wraps
import certifi
import os
from os.path import join, dirname
from dotenv import load_dotenv
import jwt
from werkzeug.utils import secure_filename
import uuid
import math
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
# Load .env
ca = certifi.where()
dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

MONGODB_URI = os.environ.get("MONGODB_URI")
DB_NAME = os.environ.get("DB_NAME")
SECRET_KEY = os.environ.get("SECRET_KEY")
TOKEN_KEY = os.environ.get("TOKEN_KEY")
client = MongoClient(MONGODB_URI, tlsCAFile=ca)
db = client[DB_NAME]

app = Flask(__name__)

def is_logged_in():
    token_receive = request.cookies.get("mytoken")
    if not token_receive:
        return False
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        return True
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return False

def get_user_info():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        user_info = db.users.find_one({'username': payload.get('id')})
        return user_info
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return None

def is_admin(user_info):
    return user_info and user_info.get('level') == 1

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            print("User not logged in, redirecting to login page")
            return redirect(url_for('page_login', msg="Anda harus login terlebih dahulu!"))
        return f(*args, **kwargs)
    return decorated_function

def is_treasurer(user_info):
    return user_info and user_info.get('level') == 3

def treasurer_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_info = get_user_info()
        if not user_info:
            return redirect(url_for('page_login', msg="Anda harus login terlebih dahulu!"))
        if not is_treasurer(user_info):
            return redirect(url_for('index', msg="Hanya bendahara yang diizinkan mengakses halaman ini!"))
        return f(*args, **kwargs)
    return decorated_function

# Fungsi validasi format kata sandi
def is_valid_password(password):
    pattern = r"^(?=.*\d)(?=.*[a-zA-Z])[0-9a-zA-Z!@#$%^&*]{8,20}$"
    return re.match(pattern, password) is not None


# Routes
@app.route('/auth_login')
def auth_login():
    token_receive = request.cookies.get(TOKEN_KEY)
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        user_info = db.users.find_one({'username': payload.get('id')})
        count_unread = db.notif.count_documents(
            {'to': payload['id'], 'from': {'$ne': payload['id']}, 'read': False})
        count_unread_payments = db.orders.count_documents(
            {'user_id': ObjectId(user_info['_id']), 'status': 'Menunggu Pembayaran'})
        data_user = {
            'username': user_info['username'],
            'profilename': user_info['profile_name'],
            'level': user_info['level'],
            'profile_icon': user_info['profile_pic_real'],
            'unread_payments': count_unread_payments
        }
        return jsonify({"result": "success", "data": data_user, "notif": count_unread})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return jsonify({"result": "fail"})

@app.route('/auth_login/<postcreator>')
def auth_login_detail(postcreator):
    token_receive = request.cookies.get(TOKEN_KEY)
    try:
        payload = jwt.decode(
            token_receive,
            SECRET_KEY,
            algorithms=["HS256"],
        )
        user_info = db.users.find_one({'username': payload.get('id')})
        if user_info['username'] == postcreator:
            return jsonify({"result": "success"})
        else:
            return jsonify({"result": "fail"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return jsonify({"result": "fail"})

@app.route('/auth_login/<commentcreator>')
def auth_login_comment(commentcreator):
    token_receive = request.cookies.get(TOKEN_KEY)
    try:
        payload = jwt.decode(
            token_receive,
            SECRET_KEY,
            algorithms=["HS256"],
        )
        user_info = db.users.find_one({'username': payload.get('id')})
        if user_info['username'] == commentcreator:
            return jsonify({"result": "success"})
        else:
            return jsonify({"result": "fail"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return jsonify({"result": "fail"})

@app.route('/login')
def page_login():
    token_receive = request.cookies.get(TOKEN_KEY)
    try:
        payload = jwt.decode(
            token_receive,
            SECRET_KEY,
            algorithms=["HS256"],
        )
        user_info = db.users.find_one({'username': payload.get('id')})
        return redirect(url_for("home", msg="Anda sudah login!"))
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return render_template('login.html')

@app.route('/register/check_dup', methods=["POST"])
def check_dup():
    username_receive = request.form.get("username_give")
    exists = bool(db.users.find_one({'username': username_receive}))
    return jsonify({"result": "success", "exists": exists})

# Ensure you have an index on the custom ID field
db.users.create_index([("user_id", ASCENDING)], unique=True)

@app.route('/register', methods=["POST"])
def register():
    nik_receive = request.form.get("nik_give")
    email_receive = request.form.get("email_give")
    password_receive = request.form.get("password_give")
    username_receive = request.form.get("username_give")
    fullname_receive = request.form.get("fullname_give")
    birthdate_receive = request.form.get("birthdate_give")
    phone_receive = request.form.get("phone_give")
    gender_receive = request.form.get("gender_give")

    # Cek apakah email sudah ada
    emailcheck = bool(db.users.find_one({'email': email_receive}))
    if emailcheck:
        return jsonify({"result": "fail", "msg": 'Maaf, email yang anda gunakan sudah terdaftar!'})

    # Validasi format kata sandi
    if not is_valid_password(password_receive):
        return jsonify({"result": "fail", "msg": "Kata sandi harus 8-20 karakter, mengandung huruf dan angka!"})

    # Buat ID pengguna berurutan
    last_user = db.users.find_one(sort=[("user_id", -1)])
    if last_user and 'user_id' in last_user:
        last_id = last_user['user_id']
        numeric_part = int(last_id[1:])
        new_id = f"A{numeric_part + 1:02d}"
    else:
        new_id = "A01"

    # Hash kata sandi
    hashed_password = generate_password_hash(password_receive)

    # Data pengguna
    data_user = {
        "user_id": new_id,
        "username": username_receive,
        "email": email_receive,
        "password": hashed_password,
        "profile_name": fullname_receive,
        "profile_pic": "",
        "profile_pic_real": "profile_pics/profile_icon.png",
        "profile_info": "",
        "blocked": False,
        "level": 2,
        "nik": nik_receive,
        "datejoin": birthdate_receive,
        "phone": phone_receive,
        "gender": gender_receive
    }

    # Masukkan data pengguna ke database
    db.users.insert_one(data_user)
    return jsonify({"result": "success", "data": email_receive})

@app.route('/login', methods=["POST"])
def login():
    email_receive = request.form["email_give"]
    password_receive = request.form["password_give"]

    result = db.users.find_one({"email": email_receive})
    if result and check_password_hash(result['password'], password_receive):
        data_user = {
            'profilename': result['profile_name'],
            'level': result['level']
        }
        if result['blocked']:
            data_block = db.blocklist.find_one({'user': result['username']})
            data_user['reasonblock'] = data_block['reason']
            data_user['userblock'] = data_block['user']
            return jsonify(
                {
                    "result": "fail",
                    "data": data_user,
                    "status": "block"
                }
            )
        else:
            payload = {
                "id": result['username'],
                "exp": datetime.utcnow() + timedelta(seconds=60 * 60 * 24),
            }
            token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")
            return jsonify(
                {
                    "result": "success",
                    "token": token,
                    "data": data_user,
                }
            )
    else:
        return jsonify(
            {
                "result": "fail",
                "msg": "Kami tidak dapat menemukan akun anda, silakan cek email dan password anda!",
                "status": "Not Found"
            }
        )
    
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/tentangkami')
def about():
    return render_template('tentangkami.html')

@app.route('/kontak')
def contact():
    return render_template('kontak.html')

@app.route('/layanan')
def services():
    return render_template('layanan.html')

@app.route('/home')
@login_required
def home():
    user_info = get_user_info()
    return render_template('dashboard_user.html', user_info=user_info)

@app.route('/dashboard')
def dashboard():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        user_login = db.users.find_one({"username": payload["id"]})
        if not user_login:
            return redirect(url_for("login", msg="Pengguna tidak ditemukan!"))
        if user_login.get('level') != 1:
            return redirect(url_for("login", msg="Anda tidak diizinkan masuk halaman dashboard!"))

        # Fetch total members with level 2 only
        total_members = db.users.count_documents({"level": 2})
        
        # Fetch total mandatory savings (Simpanan Wajib) for level 2 users
        total_mandatory_savings = db.users.aggregate([
            {"$match": {"level": 2}},  # Filter users with level 2
            {"$group": {"_id": None, "total": {"$sum": "$total_deposit_wajib"}}}
        ])
        total_mandatory_savings_value = next(total_mandatory_savings, {"total": 0})["total"]

        # Fetch total voluntary savings (Simpanan Sukarela) for level 2 users
        total_voluntary_savings = db.users.aggregate([
            {"$match": {"level": 2}},  # Filter users with level 2
            {"$group": {"_id": None, "total": {"$sum": "$total_deposit_sukarela_with_interest"}}}
        ])
        total_voluntary_savings_value = next(total_voluntary_savings, {"total": 0})["total"]

        # Fetch total loans for level 2 users
        level_2_users = db.users.find({"level": 2}, {"user_id": 1})
        level_2_user_ids = [user["user_id"] for user in level_2_users]
        total_loans = db.loans.aggregate([
            {"$match": {"user_id": {"$in": level_2_user_ids}, "status": {"$in": ["active", "pending"]}}},
            {"$group": {"_id": None, "total": {"$sum": "$loan_amount"}}}
        ])
        total_loans_value = next(total_loans, {"total": 0})["total"]

        # Fetch total income (example: sum of completed payments) for level 2 users
        total_income = db.payments.aggregate([
            {"$match": {"user_id": {"$in": level_2_user_ids}, "status": "completed"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ])
        total_income_value = next(total_income, {"total": 0})["total"]

        # Fetch recent transactions (limit to 4 for display)
        recent_transactions = []
        deposits = db.deposits.find({"user_id": {"$in": level_2_user_ids}}).sort("deposit_date", -1).limit(4)
        loans = db.loans.find({"user_id": {"$in": level_2_user_ids}}).sort("loan_date", -1).limit(4)
        payments = db.payments.find({"user_id": {"$in": level_2_user_ids}}).sort("payment_date", -1).limit(4)

        # Combine and format transactions
        for deposit in deposits:
            if "user_id" not in deposit:
                continue  # Skip if user_id is missing
            user = db.users.find_one({"user_id": deposit["user_id"]})
            recent_transactions.append({
                "date": datetime.strptime(deposit["deposit_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
                "member": user["profile_name"] if user else "Unknown",
                "type": f"Simpanan {deposit['deposit_type'].capitalize()}",
                "amount": f"Rp {deposit['deposit_amount']:,.0f}",
                "status": deposit.get("status", "completed").capitalize()
            })

        for loan in loans:
            if "user_id" not in loan:
                continue  # Skip if user_id is missing
            user = db.users.find_one({"user_id": loan["user_id"]})
            recent_transactions.append({
                "date": datetime.strptime(loan["loan_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
                "member": user["profile_name"] if user else "Unknown",
                "type": f"Pinjaman ({loan['loan_purpose']})",
                "amount": f"Rp {loan['loan_amount']:,.0f}",
                "status": loan["status"].capitalize()
            })

        for payment in payments:
            if "user_id" not in payment:
                continue  # Skip if user_id is missing
            user = db.users.find_one({"user_id": payment["user_id"]})
            recent_transactions.append({
                "date": datetime.strptime(payment["payment_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
                "member": user["profile_name"] if user else "Unknown",
                "type": "Pembayaran Angsuran",
                "amount": f"Rp {payment['amount']:,.0f}",
                "status": payment["status"].capitalize()
            })

        # Sort transactions by date and limit to 4
        recent_transactions.sort(key=lambda x: x["date"], reverse=True)
        recent_transactions = recent_transactions[:4]

        # Fetch member overview for level 2 users
        members = list(db.users.find({"level": 2}).limit(4))
        members_data = []
        for member in members:
            if "user_id" not in member:
                continue  # Skip if user_id is missing
            
            total_mandatory_savings_member = member.get("total_deposit_wajib", 0)
            total_voluntary_savings_member = member.get("total_deposit_sukarela_with_interest", 0)

            total_loans = db.loans.aggregate([
                {"$match": {"user_id": member["user_id"], "status": {"$in": ["active", "pending"]}}},
                {"$group": {"_id": None, "total": {"$sum": "$loan_amount"}}}
            ])
            total_loans_value_member = next(total_loans, {"total": 0})["total"]

            status = "Active" if not member.get("blocked", False) else "Blocked"
            if total_loans_value_member > 0:
                status = "On Loan"

            members_data.append({
                "member_id": member["user_id"],
                "name": member["profile_name"],
                "join_date": member.get("datejoin", "N/A"),
                "total_mandatory_savings": f"Rp {total_mandatory_savings_member:,.0f}",
                "total_voluntary_savings": f"Rp {total_voluntary_savings_member:,.0f}",
                "total_loans": f"Rp {total_loans_value_member:,.0f}",
                "status": status
            })

        return render_template(
            'dashboard_admin.html',
            users=user_login,
            total_members=total_members,
            total_mandatory_savings=total_mandatory_savings_value,  # Pass raw number
            total_voluntary_savings=total_voluntary_savings_value,  # Pass raw number
            total_loans=total_loans_value,  # Pass raw number
            total_income=total_income_value,  # Pass raw number
            recent_transactions=recent_transactions,
            members=members_data
        )
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return redirect(url_for("login", msg="Anda belum login!"))
                    
@app.route('/anggota', methods=['GET'])
def anggota():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        user_login = db.users.find_one({"username": payload["id"]})
        if user_login['level'] == 1:
            # Fetch only users with level 2 (exclude level 1 and level 3)
            users = list(db.users.find({"level": 2}))
            for user in users:
                user['_id'] = str(user['_id'])
            return render_template('anggota.html', users=users)
        else:
            return redirect(url_for("index", msg="Anda tidak diizinkan masuk halaman anggota!"))
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return redirect(url_for("index", msg="Anda belum login!"))
    
@app.route('/anggota/<username>', methods=['GET'])
@login_required
def member_detail(username):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index", msg="You are not authorized to access this page!"))

    # Fetch the user data
    user = db.users.find_one({"username": username})
    if not user:
        return redirect(url_for("anggota", msg="User not found!"))

    # Convert ObjectId to string for template rendering
    user['_id'] = str(user['_id'])

    # Fetch additional data if needed (e.g., total savings, loans)
    total_mandatory_savings = user.get("total_deposit_wajib", 0)
    total_voluntary_savings = user.get("total_deposit_sukarela_with_interest", 0)

    total_loans = db.loans.aggregate([
        {"$match": {"user_id": user["user_id"], "status": {"$in": ["active", "pending"]}}},
        {"$group": {"_id": None, "total": {"$sum": "$loan_amount"}}}
    ])
    total_loans_value = next(total_loans, {"total": 0})["total"]

    # Prepare user data for template
    user_data = {
        "user_id": user.get("user_id", "N/A"),
        "username": user.get("username", "N/A"),
        "profile_name": user.get("profile_name", "N/A"),
        "email": user.get("email", "N/A"),
        "nik": user.get("nik", "N/A"),
        "datejoin": user.get("datejoin", "N/A"),
        "phone": user.get("phone", "N/A"),
        "gender": user.get("gender", "N/A"),
        "profile_info": user.get("profile_info", "N/A"),
        "profile_pic_real": user.get("profile_pic_real", "profile_pics/profile_icon.png"),
        "blocked": user.get("blocked", False),
        "total_mandatory_savings": f"Rp {total_mandatory_savings:,.0f}",
        "total_voluntary_savings": f"Rp {total_voluntary_savings:,.0f}",
        "total_loans": f"Rp {total_loans_value:,.0f}",
        "status": "Active" if not user.get("blocked", False) else "Blocked"
    }

    return render_template('user_detail.html', user=user_data, user_info=user_info)

@app.route('/edit_user/<username>', methods=['GET'])
def edit_user(username):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index"))

    user = db.users.find_one({"username": username})
    return render_template("editUser.html", user=user, user_info=user_info, is_admin=True, logged_in=True)

@app.route("/update_profile", methods=["POST"])
def update_profile():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        username = payload["id"]
        fullname_receive = request.form["fullname_give"]
        email_receive = request.form["email_give"]
        job_receive = request.form["job_give"]
        phone_receive = request.form["phone_give"]
        address_receive = request.form["address_give"]
        bio_receive = request.form["bio_give"]
        new_doc = {
            "profile_name": fullname_receive,
            "email": email_receive,
            "profile_job": job_receive,
            "profile_phone": phone_receive,
            "profile_address": address_receive,
            "profile_info": bio_receive
        }
        # Validasi unggahan file
        if "file_give" in request.files:
            file = request.files["file_give"]
            if file and allowed_file(file.filename):
                if file.content_length > 5 * 1024 * 1024:  # Batas 5MB
                    return jsonify({'result': 'fail', 'msg': 'Ukuran file maksimal 5MB'})
                time = datetime.now().strftime("%m%d%H%M%S")
                filename = secure_filename(file.filename)
                extension = filename.rsplit(".", 1)[1].lower()
                file_path = f"profile_pics/profilimg-{username}-{time}.{extension}"
                file.save(os.path.join("static", file_path))
                new_doc["profile_pic"] = filename
                new_doc["profile_pic_real"] = file_path
            else:
                return jsonify({'result': 'fail', 'msg': 'Format file tidak valid. Gunakan PNG, JPG, JPEG, atau PDF'})
        
        # Cek apakah email sudah digunakan oleh pengguna lain
        user = db.users.find_one({"username": username})
        if email_receive != user['email'] and db.users.find_one({'email': email_receive}):
            return jsonify({'result': 'fail', 'msg': 'Email sudah digunakan oleh pengguna lain!'})

        db.users.update_one({"username": username}, {"$set": new_doc})  # Perbaikan: gunakan db.users
        return jsonify({"result": "success", "msg": "Profil berhasil diperbarui!"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return redirect(url_for("page_login", msg="Sesi tidak valid, silakan login kembali!"))

@app.route("/reset_pass", methods=["POST"])
def reset_pass():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        username = payload["id"]
        old_password = request.form.get("old_password_give")
        new_password = request.form.get("passnew_give")

        # Validasi input
        if not old_password or not new_password:
            return jsonify({"result": "fail", "msg": "Kata sandi lama dan baru harus diisi!"})

        user = db.users.find_one({"username": username})
        if not user:
            return jsonify({"result": "fail", "msg": "Pengguna tidak ditemukan!"})

        # Cek kata sandi lama
        if not check_password_hash(user["password"], old_password):
            return jsonify({"result": "fail", "msg": "Kata sandi lama tidak sesuai!"})

        # Cek apakah kata sandi baru sama dengan kata sandi lama
        if old_password == new_password:
            return jsonify({"result": "fail", "msg": "Kata sandi baru tidak boleh sama dengan kata sandi lama!"})

        # Validasi format kata sandi baru
        if not is_valid_password(new_password):
            return jsonify({"result": "fail", "msg": "Kata sandi baru harus 8-20 karakter, mengandung huruf dan angka!"})

        # Hash kata sandi baru
        hashed_new_password = generate_password_hash(new_password)
        db.users.update_one({"username": username}, {"$set": {"password": hashed_new_password}})
        return jsonify({"result": "success", "msg": "Kata sandi berhasil diperbarui!"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return jsonify({"result": "fail", "msg": "Sesi tidak valid, silakan login kembali!"})
    
@app.route('/delete_user/<username>', methods=['DELETE'])
def delete_user(username):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Access denied'})

    try:
        result = db.users.delete_one({"username": username})
        if result.deleted_count > 0:
            return jsonify({'result': 'success', 'msg': 'User deleted successfully'})
        else:
            return jsonify({'result': 'fail', 'msg': 'User not found'})
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': str(e)})

@app.route('/blockuser', methods=['POST'])
def blockuser():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        username = payload["id"]
        username_receive = request.form["username_give"]
        reason_receive = request.form["reason_give"]
        date_receive = request.form["date_give"]

        doc = {
            'from': username,
            'user': username_receive,
            'reason': reason_receive,
            'date': date_receive,
        }

        db.blocklist.insert_one(doc)
        db.users.update_one({"username": username_receive}, {"$set": {'blocked': True}})
        return jsonify({"result": "success", "msg": "User blocked!"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return redirect(url_for("dashboard"))

@app.route('/unblockuser', methods=['POST'])
def unblockuser():
    token_receive = request.cookies.get("mytoken")
    try:
        payload = jwt.decode(token_receive, SECRET_KEY, algorithms=["HS256"])
        username = payload["id"]
        username_receive = request.form["username_give"]

        db.blocklist.delete_one({
            'user': username_receive,
        })

        db.users.update_one({"username": username_receive}, {"$set": {'blocked': False}})
        return jsonify({"result": "success", "msg": "User unblocked!"})
    except (jwt.ExpiredSignatureError, jwt.exceptions.DecodeError):
        return redirect(url_for("dashboard"))

# Route Simpanan
# Tambahkan konfigurasi untuk simpanan
JUMLAH_SIMPANAN_WAJIB = 50000  # Jumlah tetap untuk Simpanan Wajib
FREKUENSI_SIMPANAN_WAJIB = 'bulanan'  # Frekuensi pembayaran
JUMLAH_MINIMUM_SUKARELA = 10000  # Jumlah minimum untuk Simpanan Sukarela
db.deposits.create_index([("deposit_id", ASCENDING)], unique=True)

@app.route('/simpanan', methods=['POST'])
@login_required
def simpanan():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'gagal', 'msg': 'Tidak diizinkan'})

    deposit_type = request.form.get('depositType')
    try:
        deposit_amount = float(request.form.get('depositAmount'))
    except ValueError:
        return jsonify({'result': 'gagal', 'msg': 'Jumlah simpanan harus berupa angka'})

    # Validasi jenis simpanan
    if deposit_type not in ['wajib', 'sukarela']:
        return jsonify({'result': 'gagal', 'msg': 'Jenis simpanan tidak valid'})

    # Validasi Simpanan Wajib
    if deposit_type == 'wajib':
        if deposit_amount != JUMLAH_SIMPANAN_WAJIB:
            return jsonify({
                'result': 'gagal',
                'msg': f'Simpanan Wajib harus sebesar Rp {JUMLAH_SIMPANAN_WAJIB:,}'
            })
        # Cek apakah sudah membayar Simpanan Wajib untuk periode ini
        last_deposit = db.deposits.find_one(
            {'user_id': user_info['user_id'], 'deposit_type': 'wajib'},
            sort=[('deposit_date', -1)]
        )
        if last_deposit:
            last_deposit_date = datetime.strptime(last_deposit['deposit_date'], "%Y-%m-%d %H:%M:%S")
            current_date = datetime.now()
            if FREKUENSI_SIMPANAN_WAJIB == 'bulanan' and \
               last_deposit_date.month == current_date.month and \
               last_deposit_date.year == current_date.year:
                return jsonify({
                    'result': 'gagal',
                    'msg': 'Simpanan Wajib untuk bulan ini sudah dibayar'
                })

    # Validasi Simpanan Sukarela
    if deposit_type == 'sukarela' and deposit_amount < JUMLAH_MINIMUM_SUKARELA:
        return jsonify({
            'result': 'gagal',
            'msg': f'Jumlah Simpanan Sukarela minimal Rp {JUMLAH_MINIMUM_SUKARELA:,}'
        })

    # Buat ID simpanan unik
    last_deposit = db.deposits.find_one(sort=[("deposit_id", -1)])
    if last_deposit and 'deposit_id' in last_deposit:
        last_id = last_deposit['deposit_id']
        numeric_part = int(last_id[1:])
        new_id = f"D{numeric_part + 1:03d}"
    else:
        new_id = "D001"

    # Data simpanan
    deposit_data = {
        "deposit_id": new_id,
        "user_id": user_info['user_id'],
        "deposit_type": deposit_type,
        "deposit_amount": deposit_amount,
        "deposit_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "can_withdraw": deposit_type == 'sukarela',
        "status": "active" if deposit_type == 'sukarela' else "completed"  # Add status field
    }

    # Simpan data simpanan
    db.deposits.insert_one(deposit_data)

    # Perbarui total simpanan pengguna
    deposit_field = 'total_deposit_wajib' if deposit_type == 'wajib' else 'total_deposit_sukarela'
    db.users.update_one(
        {'user_id': user_info['user_id']},
        {'$inc': {deposit_field: deposit_amount}}
    )

    # Terapkan bunga hanya untuk Simpanan Sukarela
    if deposit_type == 'sukarela':
        total_sukarela = user_info.get('total_deposit_sukarela', 0) + deposit_amount
        total_sukarela_with_interest = total_sukarela * 1.02  # Bunga 2%
        db.users.update_one(
            {'user_id': user_info['user_id']},
            {'$set': {'total_deposit_sukarela_with_interest': total_sukarela_with_interest}}
        )

    return jsonify({'result': 'success', 'msg': 'Simpanan berhasil'})

@app.route('/savings/<savings_id>', methods=['GET'])
@login_required
def view_saving(savings_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can view savings details!'}), 403

    try:
        deposit = db.deposits.find_one({'deposit_id': savings_id})
        if not deposit:
            return jsonify({'result': 'fail', 'msg': 'Savings not found!'}), 404

        user = db.users.find_one({'user_id': deposit['user_id']})
        deposit_details = {
            'savings_id': deposit['deposit_id'],
            'member_name': user['profile_name'] if user else 'Unknown',
            'deposit_type': deposit['deposit_type'].capitalize(),
            'amount': f"Rp {deposit['deposit_amount']:,.0f}",
            'date': datetime.strptime(deposit['deposit_date'], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
            'status': deposit.get('status', 'Active').capitalize(),
            'can_withdraw': deposit['can_withdraw']
        }
        return jsonify({'result': 'success', 'data': deposit_details})
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    
@app.route('/tarik_simpanan', methods=['POST'])
@login_required
def tarik_simpanan():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'gagal', 'msg': 'Tidak diizinkan'})

    try:
        withdraw_amount = float(request.form.get('withdrawAmount'))
    except ValueError:
        return jsonify({'result': 'gagal', 'msg': 'Jumlah penarikan harus berupa angka'})

    # Validate minimum withdrawal amount
    if withdraw_amount < JUMLAH_MINIMUM_SUKARELA:
        return jsonify({
            'result': 'gagal',
            'msg': f'Jumlah penarikan minimal Rp {JUMLAH_MINIMUM_SUKARELA:,}'
        })

    # Validate step (multiple of 1000)
    if withdraw_amount % 1000 != 0:
        return jsonify({
            'result': 'gagal',
            'msg': 'Jumlah penarikan harus kelipatan Rp 1.000'
        })

    # Check available balance
    total_sukarela = user_info.get('total_deposit_sukarela_with_interest', 0)
    if withdraw_amount > total_sukarela:
        return jsonify({
            'result': 'gagal',
            'msg': f'Jumlah penarikan melebihi saldo Simpanan Sukarela (Rp {total_sukarela:,.0f})'
        })
    
    # Buat ID simpanan unik
    last_withdrwa = db.withdrawals.find_one(sort=[("withdrawal_id", -1)])
    if last_withdrwa and 'withdrawal_id' in last_withdrwa:
        last_id = last_withdrwa['withdrawal_id']
        numeric_part = int(last_id[1:])
        new_id = f"W{numeric_part + 1:03d}"
    else:
        new_id = "W001"
        
    # Record withdrawal with pending status
    withdrawal_data = {

        "withdrawal_id": new_id,
        "user_id": user_info['user_id'],
        "withdraw_amount": withdraw_amount,
        "withdrawal_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "pending"  # Set status to pending
    }
    db.withdrawals.insert_one(withdrawal_data)

    return jsonify({'result': 'success', 'msg': 'Permintaan penarikan berhasil diajukan dan menunggu persetujuan'})

@app.route('/withdrawal_request', methods=['GET'])
@login_required
def withdrawal_request():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('permintaan_penarikan_user.html', user_info=user_info)

@app.route('/simpanan', methods=['GET'])
@login_required
def get_simpanan():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('simpanan_user.html', user_info=user_info)

@app.route('/get_simpanan', methods=['GET'])
@login_required
def gets_simpanan():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'gagal', 'msg': 'Tidak diizinkan'})
    
    # Ambil tanggal transaksi simpanan terbaru
    last_deposit = db.deposits.find_one(
        {'user_id': user_info['user_id']},
        sort=[('deposit_date', -1)]
    )
    last_updated = last_deposit['deposit_date'] if last_deposit else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_wajib = user_info.get('total_deposit_wajib', 0)
    total_sukarela = user_info.get('total_deposit_sukarela_with_interest', 0)
    
    return jsonify({
        'result': 'success',
        'total_wajib': total_wajib,
        'total_sukarela': total_sukarela,
        'last_updated': last_updated
    })

# Route Pinjaman
db.loans.create_index([("loan_id", ASCENDING)], unique=True)

@app.route('/pinjaman', methods=['POST'])
@login_required
def ajukan_pinjaman():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})

    try:
        loan_amount = float(request.form.get('loanAmount'))
        loan_purpose = request.form.get('loanPurpose')
        loan_term = int(request.form.get('loanTerm'))

        if loan_amount < 500000:
            return jsonify({'result': 'fail', 'msg': 'Jumlah pinjaman minimal Rp 500.000'})

        last_loan = db.loans.find_one(sort=[("loan_id", -1)])
        if last_loan and 'loan_id' in last_loan:
            last_id = last_loan['loan_id']
            numeric_part = int(last_id[1:])
            new_id = f"L{numeric_part + 1:03d}"
        else:
            new_id = "L001"

        due_date = datetime.now() + timedelta(days=loan_term * 30)

        loan_data = {
            "loan_id": new_id,
            "user_id": user_info['user_id'],
            "loan_amount": loan_amount,
            "loan_purpose": loan_purpose,
            "loan_term": loan_term,
            "loan_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "due_date": due_date.strftime("%Y-%m-%d %H:%M:%S"),
            "status": "pending",
            "amount_paid": 0
        }

        db.loans.insert_one(loan_data)
        
        # Update total_loan for the user
        db.users.update_one(
            {'user_id': user_info['user_id']},
            {'$inc': {'total_loan': loan_amount}}
        )

        return jsonify({
            'result': 'success',
            'msg': 'Pengajuan pinjaman berhasil',
            'loan_amount': loan_amount,
            'loan_purpose': loan_purpose,
            'loan_term': loan_term
        })
    except ValueError:
        return jsonify({'result': 'fail', 'msg': 'Invalid input data'})
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': str(e)})

@app.route('/pinjaman', methods=['GET'])
@login_required
def get_pinjaman():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('pinjaman_user.html', user_info=user_info)

@app.route('/get_loan_details', methods=['GET'])
@login_required
def get_loan_details():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    
    # get status pinjaman
    loans_data = db.loans.find_one({'user_id': user_info['user_id'], 'status': 'pending'})
    if loans_data:
        status = loans_data.get('status')
        return jsonify({
            'result': 'success',
            'status': status
        })
    
    loans_data = db.loans.find_one({'user_id': user_info['user_id'], 'status': 'rejected'})
    if loans_data:
        status = loans_data.get('status')
        return jsonify({
            'result': 'success',
            'status': status
        })


    loan_data = db.loans.find_one({'user_id': user_info['user_id'], 'status': 'active'})
    if loan_data:
        due_date = loan_data.get('due_date')
        loan_amount = loan_data.get('loan_amount')
        amount_paid = loan_data.get('amount_paid', 0)
        remaining_amount = loan_amount - amount_paid
        payment_progress = (amount_paid / loan_amount) * 100
        status = loan_data.get('status')

        return jsonify({
            'result': 'success',
            'due_date': due_date,
            'loan_amount': loan_amount,
            'amount_paid': amount_paid,
            'remaining_amount': remaining_amount,
            'payment_progress': payment_progress,
            'status': status
        })
    else:
        return jsonify({'result': 'fail', 'msg': 'No active loan found'})
    

# Route Pembayran/Angsuran
# Konfigurasi untuk unggahan file
UPLOAD_FOLDER = 'static/uploads/payment_proofs'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Pastikan folder unggahan ada
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/get_active_loans', methods=['GET'])
@login_required
def get_active_loans():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'gagal', 'msg': 'Sesi tidak valid. Silakan login kembali.'})

        loans = list(db.loans.find({
            'user_id': user_info['user_id'],
            'status': {'$in': ['active']}
        }))
        
        loan_options = []
        for loan in loans:
            monthly_installment = loan['loan_amount'] / loan['loan_term']
            remaining_balance = loan['loan_amount'] - loan.get('amount_paid', 0)
            loan_options.append({
                'loan_id': loan['loan_id'],
                'purpose': loan['loan_purpose'],
                'original_amount': loan['loan_amount'],
                'remaining_balance': remaining_balance,
                'monthly_installment': monthly_installment,
                'due_date': loan['due_date'],
                'term': loan['loan_term']
            })

        return jsonify({
            'result': 'success',
            'loans': loan_options
        })
    except Exception as e:
        return jsonify({
            'result': 'gagal',
            'msg': f'Terjadi kesalahan saat mengambil data pinjaman: {str(e)}'
        }), 500

@app.route('/pay_installment', methods=['POST'])
@login_required
def pay_installment():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'gagal', 'msg': 'Sesi tidak valid. Silakan login kembali.'})

        loan_id = request.form.get('loanId')
        if not loan_id:
            return jsonify({'result': 'gagal', 'msg': 'Pinjaman tidak dipilih.'})

        try:
            installment_amount = float(request.form.get('installmentAmount'))
        except (ValueError, TypeError):
            return jsonify({'result': 'gagal', 'msg': 'Jumlah angsuran harus berupa angka yang valid.'})

        # Validasi pinjaman
        loan = db.loans.find_one({'loan_id': loan_id, 'user_id': user_info['user_id'], 'status': 'active'})
        if not loan:
            return jsonify({'result': 'gagal', 'msg': 'Pinjaman tidak ditemukan atau tidak aktif.'})

        # Hitung sisa pokok dan angsuran bulanan
        remaining_balance = loan['loan_amount'] - loan.get('amount_paid', 0)
        monthly_installment = loan['loan_amount'] / loan['loan_term']
        min_installment = math.ceil(monthly_installment)  # Bulatkan ke atas, misalnya 833333.333 -> 833334

        # Validasi jumlah angsuran
        if remaining_balance <= min_installment * 2:
            # Jika sisa pokok kurang dari atau sama dengan 2x angsuran bulanan, izinkan pelunasan atau parsial
            if installment_amount < 1:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Jumlah angsuran minimal Rp 1. Anda memasukkan Rp {installment_amount:,.0f}.'
                })
            if installment_amount > remaining_balance:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Jumlah angsuran tidak boleh melebihi sisa pokok Rp {remaining_balance:,.0f}. Anda memasukkan Rp {installment_amount:,.0f}.'
                })
        else:
            # Jika sisa pokok lebih besar, terapkan angsuran minimal
            if installment_amount < min_installment:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Jumlah angsuran minimal Rp {min_installment:,.0f}. Anda memasukkan Rp {installment_amount:,.0f}. Untuk pelunasan penuh, masukkan Rp {remaining_balance:,.0f}.'
                })
            if installment_amount > remaining_balance:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Jumlah angsuran tidak boleh melebihi sisa pokok Rp {remaining_balance:,.0f}. Anda memasukkan Rp {installment_amount:,.0f}.'
                })

        # Tangani bukti pembayaran
        proof_file = request.files.get('paymentProof')
        proof_filename = None
        if proof_file and allowed_file(proof_file.filename):
            if proof_file.content_length > 5 * 1024 * 1024:
                return jsonify({'result': 'gagal', 'msg': 'Ukuran file bukti pembayaran maksimal 5MB.'})
            filename = secure_filename(f"{uuid.uuid4()}_{proof_file.filename}")
            proof_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            proof_filename = filename
        elif not proof_file:
            return jsonify({'result': 'gagal', 'msg': 'Silakan unggah bukti pembayaran.'})
        else:
            return jsonify({'result': 'gagal', 'msg': 'Format file tidak valid. Gunakan PNG, JPG, JPEG, atau PDF.'})

        # Catat pembayaran
        payment_id = f"P{datetime.now().strftime('%Y%m%d%H%M%S')}"
        payment_data = {
            'payment_id': payment_id,
            'loan_id': loan_id,
            'user_id': user_info['user_id'],
            'amount': installment_amount,
            'payment_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'proof_file': proof_filename,
            'status': 'Belum Lunas'  # Status awal
        }
        db.payments.insert_one(payment_data)

        # Perbarui jumlah terbayar
        db.loans.update_one(
            {'loan_id': loan_id},
            {'$inc': {'amount_paid': installment_amount}}
        )

        # Periksa apakah pinjaman lunas
        new_remaining_balance = remaining_balance - installment_amount
        if new_remaining_balance <= 0:
            db.loans.update_one(
                {'loan_id': loan_id},
                {'$set': {'status': 'Lunas'}}
            )
            db.users.update_one(
                {'user_id': user_info['user_id']},
                {'$inc': {'total_loan': -loan['loan_amount']}}
            )

        return jsonify({
            'result': 'success',
            'msg': 'Pembayaran angsuran berhasil diajukan dan menunggu verifikasi.'
        })
    except Exception as e:
        return jsonify({
            'result': 'gagal',
            'msg': f'Terjadi kesalahan saat memproses pembayaran: {str(e)}'
        }), 500
    
@app.route('/pembayaran', methods=['GET'])
@login_required
def pembayaran():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('pembayaran_user.html', user_info=user_info)

# Riwayat Simpanan
@app.route('/riwayat_simpanan', methods=['GET'])
@login_required
def riwayat_simpanan():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('riwayat_simpanan_user.html', user_info=user_info)

# Riwayat Pinjaman
@app.route('/riwayat_pinjaman', methods=['GET'])
@login_required
def riwayat_pinjaman():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    return render_template('riwayat_pinjaman_user.html', user_info=user_info)

@app.route('/get_deposit_history', methods=['GET'])
@login_required
def get_deposit_history():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'fail', 'msg': 'Unauthorized'})

        user_id = user_info['user_id']
        deposits = list(db.deposits.find({'user_id': user_id}).sort('deposit_date', -1))

        transactions = []
        for deposit in deposits:
            transactions.append({
                'date': deposit['deposit_date'],
                'type': f"Simpanan {deposit['deposit_type'].capitalize()}",
                'amount': deposit['deposit_amount'],
                'status': 'Sukses'
            })

        # Format dates and amounts for display
        for transaction in transactions:
            transaction['date'] = datetime.strptime(transaction['date'], "%Y-%m-%d %H:%M:%S").strftime("%d %b %Y")
            transaction['amount'] = float(transaction['amount'])

        return jsonify({
            'result': 'success',
            'transactions': transactions
        })
    except Exception as e:
        return jsonify({
            'result': 'fail',
            'msg': f'Error fetching deposit history: {str(e)}'
        }), 500
    
@app.route('/get_loan_payment_history', methods=['GET'])
@login_required
def get_loan_payment_history():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'fail', 'msg': 'Unauthorized'})

        user_id = user_info['user_id']
        transactions = []

        # Fetch loans
        loans = db.loans.find({'user_id': user_id}).sort('loan_date', -1)
        for loan in loans:
            transactions.append({
                'date': loan['loan_date'],
                'type': f"Pinjaman ({loan['loan_purpose']})",
                'amount': loan['loan_amount'],
                'status': loan['status'].capitalize()
            })

        # Fetch payments
        payments = db.payments.find({'user_id': user_id}).sort('payment_date', -1)
        for payment in payments:
            transactions.append({
                'date': payment['payment_date'],
                'type': 'Pembayaran Angsuran',
                'amount': -payment['amount'],  # Negative for payments
                'status': payment['status'].capitalize()
            })

        # Sort all transactions by date (descending) and limit to 3
        transactions.sort(key=lambda x: datetime.strptime(x['date'], "%Y-%m-%d %H:%M:%S"), reverse=True)
        transactions = transactions[:3]

        # Format dates and amounts for display
        for transaction in transactions:
            transaction['date'] = datetime.strptime(transaction['date'], "%Y-%m-%d %H:%M:%S").strftime("%d %b %Y")
            transaction['amount'] = float(transaction['amount'])

        return jsonify({
            'result': 'success',
            'transactions': transactions
        })
    except Exception as e:
        return jsonify({
            'result': 'fail',
            'msg': f'Error fetching loan and payment history: {str(e)}'
        }), 500

# Riwayarat Transaksi
@app.route('/get_transaction_history', methods=['GET'])
@login_required
def get_transaction_history():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'fail', 'msg': 'Unauthorized'})

        user_id = user_info['user_id']
        transactions = []

        # Fetch deposits
        deposits = db.deposits.find({'user_id': user_id}).sort('deposit_date', -1).limit(3)
        for deposit in deposits:
            transactions.append({
                'date': deposit['deposit_date'],
                'type': f"Simpanan {deposit['deposit_type'].capitalize()}",
                'amount': deposit['deposit_amount'],
                'status': 'Sukses'
            })

        # Fetch withdrawals
        withdrawals = db.withdrawals.find({'user_id': user_id}).sort('withdrawal_date', -1).limit(3)
        for withdrawal in withdrawals:
            transactions.append({
                'date': withdrawal['withdrawal_date'],
                'type': 'Penarikan Sukarela',
                'amount': -withdrawal['withdraw_amount'],  # Negative for withdrawals
                'status': 'Sukses'
            })

        # Fetch loans
        loans = db.loans.find({'user_id': user_id}).sort('loan_date', -1).limit(3)
        for loan in loans:
            transactions.append({
                'date': loan['loan_date'],
                'type': f"Pinjaman ({loan['loan_purpose']})",
                'amount': loan['loan_amount'],
                'status': loan['status'].capitalize()
            })

        # Fetch payments
        payments = db.payments.find({'user_id': user_id}).sort('payment_date', -1).limit(3)
        for payment in payments:
            transactions.append({
                'date': payment['payment_date'],
                'type': 'Pembayaran Angsuran',
                'amount': -payment['amount'],  # Negative for payments
                'status': payment['status'].capitalize()
            })

        # Sort all transactions by date (descending) and limit to 3
        transactions.sort(key=lambda x: datetime.strptime(x['date'], "%Y-%m-%d %H:%M:%S"), reverse=True)
        transactions = transactions[:3]

        # Format dates and amounts for display
        for transaction in transactions:
            transaction['date'] = datetime.strptime(transaction['date'], "%Y-%m-%d %H:%M:%S").strftime("%d %b %Y")
            transaction['amount'] = float(transaction['amount'])

        return jsonify({
            'result': 'success',
            'transactions': transactions
        })
    except Exception as e:
        return jsonify({
            'result': 'fail',
            'msg': f'Error fetching transaction history: {str(e)}'
        }), 500

@app.route('/update_loan_status/<loan_id>', methods=['POST'])
@login_required
def update_loan_status(loan_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can update loan status!'}), 403

    try:
        new_status = request.form.get('status').lower()
        allowed_statuses = ['pending', 'active', 'completed', 'rejected']
        if new_status not in allowed_statuses:
            return jsonify({'result': 'fail', 'msg': 'Invalid status!'}), 400

        loan = db.loans.find_one({'loan_id': loan_id})
        if not loan:
            return jsonify({'result': 'fail', 'msg': 'Loan not found!'}), 404

        # Update loan status
        db.loans.update_one(
            {'loan_id': loan_id},
            {'$set': {'status': new_status}}
        )

        # If status is 'completed', reduce user's total_loan
        if new_status == 'completed' and loan['status'] != 'completed':
            db.users.update_one(
                {'user_id': loan['user_id']},
                {'$inc': {'total_loan': -loan['loan_amount']}}
            )

        return jsonify({
            'result': 'success',
            'msg': f'Loan {loan_id} status updated to {new_status.capitalize()}!'
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500

@app.route('/loans', methods=['GET'])
@login_required
def loans():
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index", msg="You are not authorized to access the loans page!"))

    loans = list(db.loans.find())
    for loan in loans:
        user = db.users.find_one({'user_id': loan['user_id']})
        loan['member_name'] = user['profile_name'] if user else 'Unknown'
        loan['amount'] = f"Rp {loan['loan_amount']:,.0f}"
        loan['start_date'] = datetime.strptime(loan['loan_date'], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        loan['status'] = loan['status'].capitalize()

    return render_template('pinjaman_admin.html', loans=loans)

# Saldo User
@app.route('/saldo', methods=['GET'])
@login_required
def saldo():
    user_info = get_user_info()
    if not user_info:
        return jsonify({'result': 'fail', 'msg': 'Unauthorized'})
    
    total_wajib = user_info.get('total_deposit_wajib', 0)
    total_sukarela = user_info.get('total_deposit_sukarela_with_interest', 0)
    
    return render_template('saldo_user.html', user_info=user_info, total_wajib=total_wajib, total_sukarela=total_sukarela)

@app.route('/savings', methods=['GET'])
@login_required
def savings():
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index", msg="You are not authorized to access the savings page!"))

    # Fetch all savings (deposits) from the database
    deposits = list(db.deposits.find())
    for deposit in deposits:
        # Get member name from users collection
        user = db.users.find_one({'user_id': deposit['user_id']})
        deposit['member_name'] = user['profile_name'] if user else 'Unknown'
        # Format amount as currency
        deposit['amount'] = f"Rp {deposit['deposit_amount']:,.0f}"
        # Format date
        deposit['start_date'] = datetime.strptime(deposit['deposit_date'], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        # Map deposit_type to status for display (customize as needed)
        deposit['status'] = 'Active' if deposit['deposit_type'] == 'sukarela' else 'Completed'
        # Rename deposit_id to savings_id for template consistency
        deposit['savings_id'] = deposit['deposit_id']

    return render_template('simpanan_admin.html', savings=deposits)

@app.route('/update_savings_status/<savings_id>', methods=['POST'])
@login_required
def update_savings_status(savings_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can update savings status!'}), 403

    try:
        new_status = request.form.get('status').lower()
        allowed_statuses = ['pending', 'active', 'completed', 'rejected']
        if new_status not in allowed_statuses:
            return jsonify({'result': 'fail', 'msg': 'Invalid status!'}), 400

        # Update savings status in deposits collection
        deposit = db.deposits.find_one({'deposit_id': savings_id})
        if not deposit:
            return jsonify({'result': 'fail', 'msg': 'Savings not found!'}), 404

        db.deposits.update_one(
            {'deposit_id': savings_id},
            {'$set': {'status': new_status}}
        )

        # Optional: Adjust user totals if status is 'rejected' or 'completed'
        if new_status == 'rejected':
            deposit_field = 'total_deposit_wajib' if deposit['deposit_type'] == 'wajib' else 'total_deposit_sukarela'
            db.users.update_one(
                {'user_id': deposit['user_id']},
                {'$inc': {deposit_field: -deposit['deposit_amount']}}
            )
            if deposit['deposit_type'] == 'sukarela':
                db.users.update_one(
                    {'user_id': deposit['user_id']},
                    {'$set': {'total_deposit_sukarela_with_interest': 0}}
                )

        return jsonify({
            'result': 'success',
            'msg': f'Savings {savings_id} status updated to {new_status.capitalize()}!'
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500

@app.route('/delete_saving/<savings_id>', methods=['DELETE'])
@login_required
def delete_saving(savings_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can delete savings!'}), 403

    try:
        deposit = db.deposits.find_one({'deposit_id': savings_id})
        if not deposit:
            return jsonify({'result': 'fail', 'msg': 'Savings not found!'}), 404

        # Remove the deposit from the deposits collection
        result = db.deposits.delete_one({'deposit_id': savings_id})
        if result.deleted_count > 0:
            # Adjust user totals
            deposit_field = 'total_deposit_wajib' if deposit['deposit_type'] == 'wajib' else 'total_deposit_sukarela'
            db.users.update_one(
                {'user_id': deposit['user_id']},
                {'$inc': {deposit_field: -deposit['deposit_amount']}}
            )
            if deposit['deposit_type'] == 'sukarela':
                total_sukarela = db.users.find_one({'user_id': deposit['user_id']}).get('total_deposit_sukarela', 0)
                total_sukarela_with_interest = total_sukarela * 1.02  # Recalculate interest
                db.users.update_one(
                    {'user_id': deposit['user_id']},
                    {'$set': {'total_deposit_sukarela_with_interest': total_sukarela_with_interest}}
                )
            return jsonify({'result': 'success', 'msg': 'Savings deleted successfully!'})
        else:
            return jsonify({'result': 'fail', 'msg': 'Savings not found!'})
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    
# Route Profile
@app.route('/profile')
@login_required
def profile():
    user_info = get_user_info()
    if not user_info:
        return redirect(url_for('page_login', msg="Sesi tidak valid, silakan login kembali!"))
    
    # Convert ObjectId to string for template rendering
    user_info['_id'] = str(user_info['_id'])
    
    return render_template(
        'user-profile.html',
        user_info=user_info,
        user_login=user_info  # Pass user_info as user_login for edit permissions
    )

@app.route('/payments', methods=['GET'])
@login_required
def payment():
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index", msg="You are not authorized to access the payments page!"))

    # Mark overdue payments as telat
    today = datetime.now()
    db.payments.update_many(
        {
            'status': 'belum_lunas',
            'due_date': {'$lt': today.strftime("%Y-%m-%d %H:%M:%S")}
        },
        {'$set': {'status': 'telat'}}
    )

    payments = list(db.payments.find())
    for payment in payments:
        user = db.users.find_one({'user_id': payment['user_id']})
        payment['member_name'] = user['profile_name'] if user else 'Unknown'
        payment['amount'] = f"Rp {payment['amount']:,.0f}"
        payment['date'] = datetime.strptime(payment['payment_date'], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        status_map = {
            'belum_lunas': 'Belum Lunas',
            'lunas': 'Lunas',
            'telat': 'Telat'
        }
        payment['status'] = status_map.get(payment['status'], payment['status'].capitalize())

    return render_template('pembayaran_admin.html', payments=payments)

@app.route('/update_payment_status/<payment_id>', methods=['POST'])
@login_required
def update_payment_status(payment_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can update payment status!'}), 403

    try:
        new_status = request.form.get('status').lower()
        allowed_statuses = ['belum_lunas', 'lunas', 'telat']
        if new_status not in allowed_statuses:
            return jsonify({'result': 'fail', 'msg': 'Invalid status!'}), 400

        payment = db.payments.find_one({'payment_id': payment_id})
        if not payment:
            return jsonify({'result': 'fail', 'msg': 'Payment not found!'}), 404

        # If status changes from lunas to something else, reverse loan amount_paid
        if payment['status'] == 'lunas' and new_status != 'lunas':
            db.loans.update_one(
                {'loan_id': payment['loan_id']},
                {'$inc': {'amount_paid': -payment['amount']}}
            )
            # Check if loan should revert to active
            loan = db.loans.find_one({'loan_id': payment['loan_id']})
            if loan['amount_paid'] < loan['loan_amount']:
                db.loans.update_one(
                    {'loan_id': payment['loan_id']},
                    {'$set': {'status': 'active'}}
                )
        # If status changes to lunas, update loan amount_paid
        elif new_status == 'lunas' and payment['status'] != 'lunas':
            db.loans.update_one(
                {'loan_id': payment['loan_id']},
                {'$inc': {'amount_paid': payment['amount']}}
            )
            # Check if loan is fully paid
            loan = db.loans.find_one({'loan_id': payment['loan_id']})
            if loan['amount_paid'] >= loan['loan_amount']:
                db.loans.update_one(
                    {'loan_id': payment['loan_id']},
                    {'$set': {'status': 'completed'}}
                )
                db.users.update_one(
                    {'user_id': payment['user_id']},
                    {'$inc': {'total_loan': -loan['loan_amount']}}
                )

        # Update payment status
        db.payments.update_one(
            {'payment_id': payment_id},
            {'$set': {'status': new_status}}
        )

        return jsonify({
            'result': 'success',
            'msg': f'Payment {payment_id} status updated to {new_status.capitalize()}!'
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500

@app.route('/delete_payment/<payment_id>', methods=['DELETE'])
@login_required
def delete_payment(payment_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can delete payments!'}), 403

    try:
        payment = db.payments.find_one({'payment_id': payment_id})
        if not payment:
            return jsonify({'result': 'fail', 'msg': 'Payment not found!'}), 404

        # If payment was lunas, reverse loan amount_paid and check loan status
        if payment['status'] == 'lunas':
            db.loans.update_one(
                {'loan_id': payment['loan_id']},
                {'$inc': {'amount_paid': -payment['amount']}}
            )
            loan = db.loans.find_one({'loan_id': payment['loan_id']})
            if loan['amount_paid'] < loan['loan_amount']:
                db.loans.update_one(
                    {'loan_id': payment['loan_id']},
                    {'$set': {'status': 'active'}}
                )
                db.users.update_one(
                    {'user_id': payment['user_id']},
                    {'$inc': {'total_loan': loan['loan_amount']}}
                )

        # Delete proof file if exists
        if payment.get('proof_file'):
            proof_path = os.path.join(app.config['UPLOAD_FOLDER'], payment['proof_file'])
            if os.path.exists(proof_path):
                os.remove(proof_path)

        # Delete payment
        db.payments.delete_one({'payment_id': payment_id})

        return jsonify({
            'result': 'success',
            'msg': f'Payment {payment_id} deleted successfully!'
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    

@app.route('/delete_loan/<loan_id>', methods=['DELETE'])
@login_required
def delete_loan(loan_id):
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Only admins can delete loans!'}), 403

    try:
        loan = db.loans.find_one({'loan_id': loan_id})
        if not loan:
            return jsonify({'result': 'fail', 'msg': 'Loan not found!'}), 404

        # If loan is active or pending, adjust user's total_loan
        if loan['status'] in ['active', 'pending']:
            db.users.update_one(
                {'user_id': loan['user_id']},
                {'$inc': {'total_loan': -loan['loan_amount']}}
            )

        # Delete the loan
        result = db.loans.delete_one({'loan_id': loan_id})
        if result.deleted_count > 0:
            return jsonify({'result': 'success', 'msg': 'Loan deleted successfully!'})
        else:
            return jsonify({'result': 'fail', 'msg': 'Loan not found!'})
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    
@app.route('/create_treasurer', methods=['POST'])
@login_required
def create_treasurer():
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return jsonify({'result': 'fail', 'msg': 'Hanya admin yang bisa membuat bendahara!'}), 403

    nik_receive = request.form.get("nik_give")
    email_receive = request.form.get("email_give")
    password_receive = request.form.get("password_give")
    username_receive = request.form.get("username_give")
    fullname_receive = request.form.get("fullname_give")
    birthdate_receive = request.form.get("birthdate_give")
    phone_receive = request.form.get("phone_give")
    gender_receive = request.form.get("gender_give")

    # Cek apakah email sudah ada
    if db.users.find_one({'email': email_receive}):
        return jsonify({"result": "fail", "msg": 'Email sudah terdaftar!'})

    # Validasi kata sandi
    if not is_valid_password(password_receive):
        return jsonify({"result": "fail", "msg": "Kata sandi harus 8-20 karakter, mengandung huruf dan angka!"})

    # Buat ID pengguna berurutan
    last_user = db.users.find_one(sort=[("user_id", -1)])
    new_id = "B01" f"B{int(last_user['user_id'][1:]) + 1:02d}"

    # Hash kata sandi
    hashed_password = generate_password_hash(password_receive)

    # Data bendahara
    data_user = {
        "user_id": new_id,
        "username": username_receive,
        "email": email_receive,
        "password": hashed_password,
        "profile_name": fullname_receive,
        "profile_pic": "",
        "profile_pic_real": "profile_pics/profile_icon.png",
        "profile_info": "",
        "blocked": False,
        "level": 3,  # Level bendahara
        "nik": nik_receive,
        "datejoin": birthdate_receive,
        "phone": phone_receive,
        "gender": gender_receive
    }

    db.users.insert_one(data_user)
    return jsonify({"result": "success", "msg": "Akun bendahara berhasil dibuat!"})

@app.route('/create_bendahara', methods=['GET'])
@login_required
def create_bendahara():
    user_info = get_user_info()
    if not user_info or not is_admin(user_info):
        return redirect(url_for("index", msg="You are not authorized to access the create treasurer page!"))

    return render_template('create_bendahara.html', user_info=user_info)

# Route Bendahara
# Route for Treasurer Dashboard
@app.route('/bendahara')
@treasurer_required
def dashboard_bendahara():
    user_info = get_user_info()
    if not user_info:
        return redirect(url_for("page_login", msg="Anda harus login terlebih dahulu!"))

    # Fetch total mandatory savings (Simpanan Wajib) for level 2 users
    total_mandatory_savings = db.users.aggregate([
        {"$match": {"level": 2}},
        {"$group": {"_id": None, "total": {"$sum": "$total_deposit_wajib"}}}
    ])
    total_mandatory_savings_value = next(total_mandatory_savings, {"total": 0})["total"]

    # Fetch total voluntary savings (Simpanan Sukarela) for level 2 users
    total_voluntary_savings = db.users.aggregate([
        {"$match": {"level": 2}},
        {"$group": {"_id": None, "total": {"$sum": "$total_deposit_sukarela_with_interest"}}}
    ])
    total_voluntary_savings_value = next(total_voluntary_savings, {"total": 0})["total"]

    # Fetch total active loans for level 2 users
    level_2_users = db.users.find({"level": 2}, {"user_id": 1})
    level_2_user_ids = [user["user_id"] for user in level_2_users]
    total_loans = db.loans.aggregate([
        {"$match": {"user_id": {"$in": level_2_user_ids}, "status": "active"}},
        {"$group": {"_id": None, "total": {"$sum": "$loan_amount"}}}
    ])
    total_loans_value = next(total_loans, {"total": 0})["total"]

    # Fetch recent transactions (limit to 4)
    recent_transactions = []
    deposits = db.deposits.find({"user_id": {"$in": level_2_user_ids}}).sort("deposit_date", -1).limit(4)
    loans = db.loans.find({"user_id": {"$in": level_2_user_ids}}).sort("loan_date", -1).limit(4)
    payments = db.payments.find({"user_id": {"$in": level_2_user_ids}}).sort("payment_date", -1).limit(4)

    for deposit in deposits:
        if "user_id" not in deposit:
            continue
        user = db.users.find_one({"user_id": deposit["user_id"]})
        recent_transactions.append({
            "date": datetime.strptime(deposit["deposit_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
            "member": user["profile_name"] if user else "Unknown",
            "type": f"Simpanan {deposit['deposit_type'].capitalize()}",
            "amount": f"Rp {deposit['deposit_amount']:,.0f}",
            "status": deposit.get("status", "completed").capitalize()
        })

    for loan in loans:
        if "user_id" not in loan:
            continue
        user = db.users.find_one({"user_id": loan["user_id"]})
        recent_transactions.append({
            "date": datetime.strptime(loan["loan_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
            "member": user["profile_name"] if user else "Unknown",
            "type": f"Pinjaman ({loan['loan_purpose']})",
            "amount": f"Rp {loan['loan_amount']:,.0f}",
            "status": loan["status"].capitalize()
        })

    for payment in payments:
        if "user_id" not in payment:
            continue
        user = db.users.find_one({"user_id": payment["user_id"]})
        recent_transactions.append({
            "date": datetime.strptime(payment["payment_date"], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d"),
            "member": user["profile_name"] if user else "Unknown",
            "type": "Pembayaran Angsuran",
            "amount": f"Rp {payment['amount']:,.0f}",
            "status": payment["status"].capitalize()
        })

    recent_transactions.sort(key=lambda x: x["date"], reverse=True)
    recent_transactions = recent_transactions[:4]

    return render_template(
        'dashboard_bendahara.html',
        user_info=user_info,
        total_mandatory_savings=f"Rp {total_mandatory_savings_value:,.0f}",
        total_voluntary_savings=f"Rp {total_voluntary_savings_value:,.0f}",
        total_loans=f"Rp {total_loans_value:,.0f}",
        transactions=recent_transactions
    )

@app.route('/record_transaction', methods=['POST'])
@treasurer_required
def record_transaction():
    try:
        user_info = get_user_info()
        if not user_info:
            return jsonify({'result': 'gagal', 'msg': 'Sesi tidak valid. Silakan login kembali.'}), 401

        # Get form data
        username = request.form.get('username')
        transaction_type = request.form.get('transaction_type')
        amount = float(request.form.get('amount', 0))
        loan_id = request.form.get('loanId') if transaction_type == 'payment_loan' else None
        proof_file = request.files.get('paymentProof') if transaction_type == 'payment_loan' else None

        # Validate username
        user = db.users.find_one({'username': username, 'level': 2})
        if not user:
            return jsonify({'result': 'gagal', 'msg': 'Username anggota tidak ditemukan!'}), 400

        # Validate transaction type
        if transaction_type not in ['wajib', 'sukarela', 'payment_loan']:
            return jsonify({'result': 'gagal', 'msg': 'Jenis transaksi tidak valid!'}), 400

        # Validate amount
        if amount <= 0:
            return jsonify({'result': 'gagal', 'msg': 'Jumlah harus lebih dari 0!'}), 400

        # Handle deposits (wajib or sukarela)
        if transaction_type in ['wajib', 'sukarela']:
            # Validate Simpanan Wajib
            if transaction_type == 'wajib' and amount != JUMLAH_SIMPANAN_WAJIB:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Simpanan Wajib harus sebesar Rp {JUMLAH_SIMPANAN_WAJIB:,}'
                }), 400

            # Check for existing mandatory deposit in the current month
            if transaction_type == 'wajib':
                last_deposit = db.deposits.find_one(
                    {'user_id': user['user_id'], 'deposit_type': 'wajib'},
                    sort=[('deposit_date', -1)]
                )
                if last_deposit:
                    last_deposit_date = datetime.strptime(last_deposit['deposit_date'], "%Y-%m-%d %H:%M:%S")
                    current_date = datetime.now()
                    if last_deposit_date.month == current_date.month and last_deposit_date.year == current_date.year:
                        return jsonify({
                            'result': 'gagal',
                            'msg': 'Simpanan Wajib untuk bulan ini sudah dibayar'
                        }), 400

            # Validate Simpanan Sukarela
            if transaction_type == 'sukarela' and amount < JUMLAH_MINIMUM_SUKARELA:
                return jsonify({
                    'result': 'gagal',
                    'msg': f'Jumlah Simpanan Sukarela minimal Rp {JUMLAH_MINIMUM_SUKARELA:,}'
                }), 400

            # Generate unique deposit ID
            last_deposit = db.deposits.find_one(sort=[("deposit_id", -1)])
            if last_deposit and 'deposit_id' in last_deposit:
                last_id = last_deposit['deposit_id']
                numeric_part = int(last_id[1:])
                new_id = f"D{numeric_part + 1:03d}"
            else:
                new_id = "D001"

            # Create deposit record
            deposit_data = {
                "deposit_id": new_id,
                "user_id": user['user_id'],
                "deposit_type": transaction_type,
                "deposit_amount": amount,
                "deposit_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "can_withdraw": transaction_type == 'sukarela',
                "status": "completed"  # Treasurer deposits are immediately completed
            }
            db.deposits.insert_one(deposit_data)

            # Update user's savings
            deposit_field = 'total_deposit_wajib' if transaction_type == 'wajib' else 'total_deposit_sukarela'
            db.users.update_one(
                {'user_id': user['user_id']},
                {'$inc': {deposit_field: amount}}
            )

            # Apply interest for voluntary savings
            if transaction_type == 'sukarela':
                total_sukarela = user.get('total_deposit_sukarela', 0) + amount
                total_sukarela_with_interest = total_sukarela * 1.02  # 2% interest
                db.users.update_one(
                    {'user_id': user['user_id']},
                    {'$set': {'total_deposit_sukarela_with_interest': total_sukarela_with_interest}}
                )

            return jsonify({
                'result': 'success',
                'msg': f'Simpanan {transaction_type.capitalize()} berhasil dicatat untuk {user["profile_name"]}!'
            })

        # Handle loan payment
        if transaction_type == 'payment_loan':
            if not loan_id:
                return jsonify({'result': 'gagal', 'msg': 'ID Pinjaman harus dipilih!'}), 400

            # Validate loan
            loan = db.loans.find_one({'loan_id': loan_id, 'user_id': user['user_id'], 'status': 'active'})
            if not loan:
                return jsonify({'result': 'gagal', 'msg': 'Pinjaman tidak ditemukan atau tidak aktif!'}), 400

            # Validate payment amount
            remaining_balance = loan['loan_amount'] - loan.get('amount_paid', 0)
            monthly_installment = loan['loan_amount'] / loan['loan_term']
            min_installment = math.ceil(monthly_installment)

            if remaining_balance <= min_installment * 2:
                if amount < 1:
                    return jsonify({
                        'result': 'gagal',
                        'msg': f'Jumlah angsuran minimal Rp 1. Anda memasukkan Rp {amount:,.0f}.'
                    }), 400
                if amount > remaining_balance:
                    return jsonify({
                        'result': 'gagal',
                        'msg': f'Jumlah angsuran tidak boleh melebihi sisa pokok Rp {remaining_balance:,.0f}.'
                    }), 400
            else:
                if amount < min_installment:
                    return jsonify({
                        'result': 'gagal',
                        'msg': f'Jumlah angsuran minimal Rp {min_installment:,.0f}.'
                    }), 400
                if amount > remaining_balance:
                    return jsonify({
                        'result': 'gagal',
                        'msg': f'Jumlah angsuran tidak boleh melebihi sisa pokok Rp {remaining_balance:,.0f}.'
                    }), 400

            # Handle proof file
            proof_filename = None
            if proof_file and allowed_file(proof_file.filename):
                if proof_file.content_length > 5 * 1024 * 1024:
                    return jsonify({'result': 'gagal', 'msg': 'Ukuran file bukti pembayaran maksimal 5MB.'}), 400
                filename = secure_filename(f"{uuid.uuid4()}_{proof_file.filename}")
                proof_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                proof_filename = filename
            else:
                return jsonify({'result': 'gagal', 'msg': 'Silakan unggah bukti pembayaran yang valid!'}), 400

            # Create payment record
            payment_id = f"P{datetime.now().strftime('%Y%m%d%H%M%S')}"
            payment_data = {
                'payment_id': payment_id,
                'loan_id': loan_id,
                'user_id': user['user_id'],
                'amount': amount,
                'payment_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'proof_file': proof_filename,
                'status': 'lunas'  # Treasurer payments are marked as completed
            }
            db.payments.insert_one(payment_data)

            # Update loan's amount_paid
            db.loans.update_one(
                {'loan_id': loan_id},
                {'$inc': {'amount_paid': amount}}
            )

            # Check if loan is fully paid
            new_remaining_balance = remaining_balance - amount
            if new_remaining_balance <= 0:
                db.loans.update_one(
                    {'loan_id': loan_id},
                    {'$set': {'status': 'completed'}}
                )
                db.users.update_one(
                    {'user_id': user['user_id']},
                    {'$inc': {'total_loan': -loan['loan_amount']}}
                )

            return jsonify({
                'result': 'success',
                'msg': f'Pembayaran angsuran sebesar Rp {amount:,.0f} berhasil dicatat untuk {user["profile_name"]}!'
            })

    except ValueError:
        return jsonify({'result': 'gagal', 'msg': 'Jumlah harus berupa angka yang valid!'}), 400
    except Exception as e:
        return jsonify({'result': 'gagal', 'msg': f'Terjadi kesalahan: {str(e)}'}), 500

# Route for Withdrawal Requests
@app.route('/withdrawal_requests', methods=['GET'])
@treasurer_required
def withdrawal_requests():
    try:
        withdrawals = list(db.withdrawals.find({"status": {"$in": ["pending", "approved", "rejected"]}}))
        withdrawal_data = []
        for withdrawal in withdrawals:
            user = db.users.find_one({"user_id": withdrawal["user_id"]})
            withdrawal_data.append({
                "withdrawal_id": withdrawal["withdrawal_id"],
                "user_id": withdrawal["user_id"],
                "member_name": user["profile_name"] if user else "Unknown",
                "withdraw_amount": withdrawal["withdraw_amount"],
                "withdrawal_date": withdrawal["withdrawal_date"],
                "status": withdrawal.get("status", "pending").capitalize()
            })
        return jsonify(withdrawal_data)
    except Exception as e:
        return jsonify({"result": "fail", "msg": f"Error: {str(e)}"}), 500

# Route to Approve or Reject Withdrawal
@app.route('/approve_withdrawal/<withdrawal_id>', methods=['POST'])
@treasurer_required
def approve_withdrawal(withdrawal_id):
    try:
        new_status = request.form.get('status').lower()
        if new_status not in ['approved', 'rejected']:
            return jsonify({'result': 'fail', 'msg': 'Invalid status!'}), 400

        withdrawal = db.withdrawals.find_one({'withdrawal_id': withdrawal_id})
        if not withdrawal:
            return jsonify({'result': 'fail', 'msg': 'Withdrawal not found!'}), 404

        # Validate available balance for approval
        if new_status == 'approved':
            user = db.users.find_one({'user_id': withdrawal['user_id']})
            total_sukarela = user.get('total_deposit_sukarela_with_interest', 0)
            if withdrawal['withdraw_amount'] > total_sukarela:
                return jsonify({
                    'result': 'fail',
                    'msg': f'Jumlah penarikan melebihi saldo Simpanan Sukarela (Rp {total_sukarela:,.0f})'
                }), 400

        # Update withdrawal status
        db.withdrawals.update_one(
            {'withdrawal_id': withdrawal_id},
            {'$set': {'status': new_status}}
        )

        # If approved, deduct from user's voluntary savings
        if new_status == 'approved':
            db.users.update_one(
                {'user_id': withdrawal['user_id']},
                {
                    '$inc': {
                        'total_deposit_sukarela': -withdrawal['withdraw_amount'],
                        'total_deposit_sukarela_with_interest': -(withdrawal['withdraw_amount'] * 1.02)
                    }
                }
            )

        return jsonify({
            'result': 'success',
            'msg': f'Withdrawal {withdrawal_id} {new_status.capitalize()}!'
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    
# Route for Treasurer Report
@app.route('/treasurer_report', methods=['GET'])
@treasurer_required
def treasurer_report():
    import datetime
    try:
        # Get pagination parameters
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 8))
        skip = (page - 1) * per_page

        # Fetch deposits
        total_deposits_count = db.deposits.count_documents({})
        deposits = list(db.deposits.find().skip(skip).limit(per_page))
        deposit_data = []
        total_deposits = sum(deposit['deposit_amount'] for deposit in db.deposits.find())  # Full dataset for total
        for deposit in deposits:
            user = db.users.find_one({'user_id': deposit['user_id']})
            deposit_amount = deposit['deposit_amount']
            deposit_data.append({
                'user_id': deposit['user_id'],
                'profile_name': user['profile_name'] if user else 'Unknown',
                'deposit_type': deposit['deposit_type'],
                'deposit_amount': deposit_amount,
                'deposit_date': deposit['deposit_date']
            })

        # Fetch payments
        total_payments_count = db.payments.count_documents({})
        payments = list(db.payments.find().skip(skip).limit(per_page))
        payment_data = []
        total_payments = sum(payment['amount'] for payment in db.payments.find())  # Full dataset for total
        for payment in payments:
            user = db.users.find_one({'user_id': payment['user_id']})
            payment_amount = payment['amount']
            payment_data.append({
                'user_id': payment['user_id'],
                'profile_name': user['profile_name'] if user else 'Unknown',
                'amount': payment_amount,
                'status': payment['status'],
                'payment_date': payment['payment_date']
            })

        return jsonify({
            'deposits': deposit_data,
            'payments': payment_data,
            'total_deposits': total_deposits,
            'total_payments': total_payments,
            'pagination': {
                'deposits': {
                    'total': total_deposits_count,
                    'pages': (total_deposits_count + per_page - 1) // per_page,
                    'current_page': page,
                    'per_page': per_page
                },
                'payments': {
                    'total': total_payments_count,
                    'pages': (total_payments_count + per_page - 1) // per_page,
                    'current_page': page,
                    'per_page': per_page
                }
            }
        })
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
# New route to export savings report as XLSX
@app.route('/export_savings_xlsx', methods=['GET'])
@treasurer_required
def export_savings_xlsx():
    import datetime
    try:
        # Fetch data from treasurer_report
        deposits = list(db.deposits.find())
        total_deposits = sum(deposit['deposit_amount'] for deposit in deposits)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Simpanan"

        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Metadata
        ws['A1'] = "Koperasi Artha Sejati"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_center
        ws.merge_cells('A1:E1')

        ws['A2'] = "Laporan Simpanan"
        ws['A2'].font = title_font
        ws['A2'].alignment = align_center
        ws.merge_cells('A2:E2')

        current_date = datetime.datetime.now().strftime("%d/%m/%Y")
        ws['A3'] = f"Tanggal Laporan: {current_date}"
        ws['A3'].font = Font(size=12)
        ws['A3'].alignment = align_center
        ws.merge_cells('A3:E3')

        # Headers
        headers = ['ID Anggota', 'Nama', 'Jenis Simpanan', 'Jumlah', 'Tanggal']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border
            cell.fill = header_fill

        # Data
        row = 6
        for deposit in deposits:
            user = db.users.find_one({'user_id': deposit['user_id']})
            profile_name = user['profile_name'] if user else 'Unknown'
            deposit_type = 'Wajib' if deposit['deposit_type'] == 'wajib' else 'Sukarela'
            amount = deposit['deposit_amount']
            date = datetime.datetime.strptime(deposit['deposit_date'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")

            ws.cell(row=row, column=1).value = deposit['user_id']
            ws.cell(row=row, column=2).value = profile_name
            ws.cell(row=row, column=3).value = deposit_type
            ws.cell(row=row, column=4).value = amount
            ws.cell(row=row, column=4).number_format = '"Rp "#,##0'
            ws.cell(row=row, column=5).value = date

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_left if col in [1, 2, 3] else align_center

            row += 1

        # Summary row
        ws.cell(row=row, column=4).value = total_deposits
        ws.cell(row=row, column=4).number_format = '"Rp "#,##0'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).alignment = align_center
        ws.cell(row=row, column=1).value = "Total Simpanan"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = align_left
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(f'A{row}:C{row}')

        # Adjust column widths
        column_widths = [15, 25, 15, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='laporan_simpanan_artha_sejati.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500

# New route to export loans report as XLSX
@app.route('/export_loans_xlsx', methods=['GET'])
@treasurer_required
def export_loans_xlsx():
    try:
        # Fetch data from payments
        payments = list(db.payments.find())
        total_payments = sum(payment['amount'] for payment in payments)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pembayaran Pinjaman"

        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Metadata
        ws['A1'] = "Koperasi Artha Sejati"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_center
        ws.merge_cells('A1:E1')

        ws['A2'] = "Laporan Pembayaran Pinjaman"
        ws['A2'].font = title_font
        ws['A2'].alignment = align_center
        ws.merge_cells('A2:E2')

        current_date = datetime.datetime.now().strftime("%d/%m/%Y")
        ws['A3'] = f"Tanggal Laporan: {current_date}"
        ws['A3'].font = Font(size=12)
        ws['A3'].alignment = align_center
        ws.merge_cells('A3:E3')

        # Headers
        headers = ['ID Anggota', 'Nama', 'Jumlah', 'Status', 'Tanggal']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border
            cell.fill = header_fill

        # Data
        row = 6
        for payment in payments:
            user = db.users.find_one({'user_id': payment['user_id']})
            profile_name = user['profile_name'] if user else 'Unknown'
            amount = payment['amount']
            status = payment['status'].capitalize()
            date = datetime.datetime.strptime(payment['payment_date'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")

            ws.cell(row=row, column=1).value = payment['user_id']
            ws.cell(row=row, column=2).value = profile_name
            ws.cell(row=row, column=3).value = amount
            ws.cell(row=row, column=3).number_format = '"Rp "#,##0'
            ws.cell(row=row, column=4).value = status
            ws.cell(row=row, column=5).value = date

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_left if col in [1, 2, 4] else align_center

            row += 1

        # Summary row
        ws.cell(row=row, column=3).value = total_payments
        ws.cell(row=row, column=3).number_format = '"Rp "#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = align_center
        ws.cell(row=row, column=1).value = "Total Pembayaran Pinjaman"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = align_left
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(f'A{row}:B{row}')

        # Adjust column widths
        column_widths = [15, 25, 20, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='laporan_pembayaran_pinjaman_artha_sejati.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'result': 'fail', 'msg': f'Error: {str(e)}'}), 500
    
# Route to Validate Username
@app.route('/validate_username', methods=['POST'])
def validate_username():
    username = request.form.get('username')
    user = db.users.find_one({'username': username, 'level': 2})
    if user:
        return jsonify({'result': 'success', 'user_id': user['user_id']})
    return jsonify({'result': 'fail', 'msg': 'Username tidak ditemukan atau bukan anggota!'})

@app.route('/permintaan_penarikan', methods=['GET'])
@treasurer_required
def permintaan_penarikan():
    return render_template('penarikan_bendahara.html')

@app.route('/get_loans', methods=['GET'])
@treasurer_required
def get_loans():
    user_id = request.args.get('user_id')
    loans = list(db.loans.find({'user_id': user_id, 'status': 'active'}, {
        'loan_id': 1, 'purpose': 1, 'remaining_balance': {'$subtract': ['$loan_amount', '$amount_paid']}, '_id': 0
    }))
    print(f"Loans found for user_id {user_id}: {loans}")  # Debugging
    return jsonify({'result': 'success', 'loans': loans})

    
# Route Logout
@app.route('/logout')
def logout():
    response = redirect(url_for('login'))
    response.delete_cookie("mytoken")
    return response


if __name__ == '__main__':
    app.run('0.0.0.0', port=5000, debug=True)